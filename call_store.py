from ind_poker import ind_poker_game
import openpyxl as xl
wb = xl.load_workbook("operationD.xlsx")
seed_value = -1
game_no = 0
for sheet_number in range(1, 51):

    pok_string = f"Data {sheet_number}"
    sheet = wb.create_sheet(pok_string, sheet_number)

    cell1 = sheet.cell(1, 1)
    cell1.value = "Game no."
    cell2 = sheet.cell(1, 2)
    cell2.value = "Number of rounds"
    cell3 = sheet.cell(1, 3)
    cell3.value = "Caller"
    cell4 = sheet.cell(1, 4)
    cell4.value = "Player 1 cards"
    cell5 = sheet.cell(1, 5)
    cell5.value = "Player 2 cards"
    cell6 = sheet.cell(1, 6)
    cell6.value = "Player 1 total"
    cell7 = sheet.cell(1, 7)
    cell7.value = "Player 2 total"
    cell8 = sheet.cell(1, 8)
    cell8.value = "Result"

    for row in range(3, 103):
        game_no += 1
        seed_value += 1
        for column in range(2, 9):
            needed = ind_poker_game(seed_value)
            cell = sheet.cell(row, 1)
            cell.value = game_no
            cell = sheet.cell(row, column)
            cell.value = needed[column - 2]



wb.save("indian_poker_collection.xlsx")